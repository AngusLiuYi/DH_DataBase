import openpyxl
from collections import Counter

SRC = r'D:\AI\DH_DataBase\raw\fae-qa\华东工单提取仅技术服务&高价值FAQ_20260820.xlsx'
OUT = r'D:\AI\DH_DataBase\raw\fae-qa\华东工单FAQ_技术服务高价值_语料_20260820.md'

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['服务工单']
hdr = [c.value for c in ws[1]]
col = {n: i + 1 for i, n in enumerate(hdr)}

def g(row, name):
    v = row[col[name] - 1]
    return str(v).strip() if v is not None else ''

prod_c, typ_c, solve_c, val_c = Counter(), Counter(), Counter(), Counter()
total = 0
blocks = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    total += 1
    prod = g(row, '执行器产品线')
    typ = g(row, '问题类型')
    solve = g(row, '解决方式')
    val = g(row, '参考价值')
    prod_c[prod] += 1
    typ_c[typ] += 1
    solve_c[solve] += 1
    val_c[val] += 1
    no = g(row, '服务单号')
    cust = g(row, '服务客户名称')
    model = g(row, '执行器型号1')
    part = g(row, '本次问题部件')
    phen = g(row, '问题现象描述')
    steps = g(row, '分析步骤')
    concl = g(row, '结论及处理')
    blocks.append(
        f"## 问题：{phen}\n\n"
        f"- **问题类型**：{typ}\n"
        f"- **产品线**：{prod}\n"
        f"- **型号**：{model}\n"
        f"- **客户**：{cust}\n"
        f"- **服务单号**：{no}\n"
        f"- **问题部件**：{part}\n\n"
        f"**现象描述**：{phen}\n\n"
        f"**排查/分析步骤**：{steps}\n\n"
        f"**结论及处理**：{concl}\n\n"
        f"---\n"
    )

header = (
    "# 华东工单 FAQ 语料（技术服务·高价值）\n\n"
    f"- 来源文件：`华东工单提取仅技术服务&高价值FAQ_20260820.xlsx`\n"
    f"- 条目数：**{total}**\n"
    "- 生成日期：2026-08-20\n"
    "- 用途：供产品知识库（RAGFlow / FastGPT / Dify）导入投喂。每条为一问一答式 FAQ，含现象 / 排查步骤 / 结论及元数据。\n\n"
    "## 统计概览\n\n"
    f"- 解决方式分布：{dict(solve_c)}\n"
    f"- 参考价值分布：{dict(val_c)}\n"
    f"- 产品线 TOP：{prod_c.most_common(10)}\n"
    f"- 问题类型 TOP：{typ_c.most_common(15)}\n\n"
    "---\n\n"
)

with open(OUT, 'w', encoding='utf-8') as f:
    f.write(header)
    f.write('\n'.join(blocks))

print('OK 已生成:', OUT)
print('条数:', total)
print('解决方式分布:', dict(solve_c))
print('参考价值分布:', dict(val_c))
