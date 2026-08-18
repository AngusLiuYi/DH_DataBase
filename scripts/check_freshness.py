#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
check_freshness.py — DH_DataBase 源文件新鲜度巡检（基于 SHA-256 内容指纹）

为什么需要它：
  光记「文件名 + 版本」抓不到「文件名没变、内容被改了」的情况。
  本工具给每个源文件算一个 SHA-256 指纹（内容哈希），存进 MANIFEST.xlsx，
  之后随时比对就能知道哪些文件被改过 / 新增 / 缺失，不用你专门来告诉我。

用法：
  python check_freshness.py baseline    # 全量重新扫描 raw/，把 sha256/mtime/size 写回 MANIFEST.xlsx
  python check_freshness.py check       # 扫描当前状态，与 MANIFEST 指纹比对，输出变更报告
  python check_freshness.py maybe       # 看上次检查时间戳：超 interval 天就自动跑 check，否则提示未超期

约定：
  - MANIFEST 位于 raw/MANIFEST.xlsx（Excel）
  - 指纹列：sha256 / mtime / size（在「适配产品线」列之后新增）
  - 不修改任何源文件；只读 raw/ 计算哈希，仅在 MANIFEST.xlsx 与 state 文件写元数据
  - 状态机：state 文件 scripts/.freshness_state.json 记录 last_check 与 interval_days

注意：baseline / check 每次运行都会更新 last_check 时间戳。
"""
import os
import sys
import json
import hashlib
import datetime

try:
    import openpyxl
except ImportError:
    sys.stderr.write("缺少 openpyxl，请先安装：pip install openpyxl\n")
    sys.exit(2)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # D:\AI\DH_DataBase
RAW = os.path.join(ROOT, "raw")
MANIFEST = os.path.join(RAW, "MANIFEST.xlsx")
STATE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".freshness_state.json")
HASH_COLS = ["sha256", "mtime", "size"]
BASE_COLS = ["文件名", "类型", "版本", "状态", "最后投喂日", "适配产品线"]

# 扫描时跳过的文件 / 目录
SKIP_FILES = {".gitkeep", "MANIFEST.xlsx", "MANIFEST.csv"}
SKIP_DIR_PREFIXES = (".")  # 隐藏目录（如 .freshness 之类）不扫

EXT_TYPE = {
    ".md": "文档", ".pdf": "手册", ".docx": "手册", ".doc": "手册",
    ".xlsx": "表格", ".xls": "表格", ".pptx": "演示", ".ppt": "演示",
    ".mp4": "视频", ".mov": "视频", ".png": "图片", ".jpg": "图片",
    ".jpeg": "图片", ".txt": "文档",
}


def sha256_of(path, chunk=1 << 16):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for blk in iter(lambda: f.read(chunk), b""):
            h.update(blk)
    return h.hexdigest()


def scan_raw():
    """返回 [(relpath, basename, sha256, mtime_iso, size), ...]"""
    out = []
    for dirpath, dirnames, filenames in os.walk(RAW):
        # 过滤隐藏目录
        dirnames[:] = [d for d in dirnames if not d.startswith(SKIP_DIR_PREFIXES)]
        for fn in filenames:
            if fn in SKIP_FILES:
                continue
            full = os.path.join(dirpath, fn)
            rel = os.path.relpath(full, RAW).replace(os.sep, "/")
            try:
                st = os.stat(full)
                out.append((
                    rel, fn,
                    sha256_of(full),
                    datetime.datetime.fromtimestamp(st.st_mtime).isoformat(timespec="seconds"),
                    st.st_size,
                ))
            except OSError as e:
                sys.stderr.write(f"无法读取 {rel}: {e}\n")
    out.sort(key=lambda r: r[0])
    return out


def load_manifest():
    """读取 MANIFEST.xlsx → (header_list, [row_dict,...])。文件不存在返回 ([], [])。"""
    if not os.path.exists(MANIFEST):
        return [], []
    wb = openpyxl.load_workbook(MANIFEST, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    header = [str(h) if h is not None else "" for h in rows[0]]
    data = []
    for r in rows[1:]:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        d = {}
        for i, h in enumerate(header):
            d[h] = r[i] if i < len(r) else None
        data.append(d)
    return header, data


def ensure_cols(header):
    """返回补齐 HASH_COLS 后的 header（HASH_COLS 追加到末尾或紧跟已有列之后）。"""
    h = list(header)
    for c in HASH_COLS:
        if c not in h:
            h.append(c)
    # 保证 BASE_COLS 顺序在前（若 header 已有这些列则保持原位，缺失则补到前面）
    ordered = [c for c in BASE_COLS if c in h]
    ordered += [c for c in h if c not in BASE_COLS and c not in HASH_COLS]
    ordered += [c for c in HASH_COLS if c in h]
    return ordered


def save_manifest(header, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MANIFEST"
    ws.append(header)
    for d in data:
        ws.append([d.get(c) for c in header])
    wb.save(MANIFEST)
    wb.close()


def load_state():
    if os.path.exists(STATE):
        try:
            with open(STATE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"last_check": None, "interval_days": 7}


def save_state(state):
    os.makedirs(os.path.dirname(STATE), exist_ok=True)
    with open(STATE, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


def stamp_now():
    return datetime.datetime.now().isoformat(timespec="seconds")


def cmd_baseline():
    header, data = load_manifest()
    if not header:
        header = list(BASE_COLS) + HASH_COLS
    scanned = scan_raw()
    by_name = {s[1]: s for s in scanned}
    seen = set()

    for d in data:
        name = d.get("文件名")
        if name in by_name:
            _, bn, h, mt, sz = by_name[name]
            d["sha256"] = h
            d["mtime"] = mt
            d["size"] = sz
            if d.get("状态") in (None, "", "缺失(磁盘无此文件)"):
                d["状态"] = "在库"
            seen.add(name)
        else:
            # 磁盘上已无此文件
            if d.get("状态") != "缺失(磁盘无此文件)":
                d["状态"] = "缺失(磁盘无此文件)"

    # 磁盘上有、MANIFEST 里没有的 → 新增待归类
    for s in scanned:
        rel, bn, h, mt, sz = s
        if bn in seen:
            continue
        ext = os.path.splitext(bn)[1].lower()
        data.append({
            "文件名": bn, "类型": EXT_TYPE.get(ext, "其他"), "版本": "",
            "状态": "新增待归类", "最后投喂日": "", "适配产品线": "",
            "sha256": h, "mtime": mt, "size": sz,
        })

    header = ensure_cols(header)
    save_manifest(header, data)
    state = load_state()
    state["last_check"] = stamp_now()
    save_state(state)
    print(f"[baseline] 已扫描 {len(scanned)} 个文件，写回 {MANIFEST}")
    print(f"[baseline] MANIFEST 现有 {len(data)} 行（含缺失/新增）")
    print(f"[baseline] last_check = {state['last_check']}")


def cmd_check():
    header, data = load_manifest()
    base = {d.get("文件名"): d for d in data if d.get("文件名")}
    scanned = scan_raw()
    cur_names = {s[1] for s in scanned}
    base_names = set(base.keys())

    changed, unchanged, new, missing = [], [], [], []
    for s in scanned:
        rel, bn, h, mt, sz = s
        if bn in base:
            if str(base[bn].get("sha256", "")) == h:
                unchanged.append((rel, bn))
            else:
                changed.append((rel, bn, base[bn].get("sha256", ""), h))
        else:
            new.append((rel, bn))
    for bn, d in base.items():
        if bn not in cur_names:
            missing.append((bn, d.get("状态")))

    print("=" * 60)
    print(f"源文件新鲜度检查  @ {stamp_now()}")
    print("=" * 60)
    if changed:
        print(f"\n🔴 内容变更（{len(changed)}）：文件名相同但内容哈希变了")
        for rel, bn, old, newh in changed:
            print(f"   - {rel}")
            print(f"       old: {old}")
            print(f"       new: {newh}")
    if new:
        print(f"\n🟡 新增（{len(new)}）：磁盘有、MANIFEST 未登记")
        for rel, bn in new:
            print(f"   + {rel}")
    if missing:
        print(f"\n⚪ 缺失（{len(missing)}）：MANIFEST 登记、磁盘已无")
        for bn, st in missing:
            print(f"   - {bn}  [{st}]")
    if unchanged:
        print(f"\n🟢 未变（{len(unchanged)}）")
    print("\n" + "-" * 60)
    print(f"汇总：未变 {len(unchanged)} | 变更 {len(changed)} | 新增 {len(new)} | 缺失 {len(missing)}")
    print("-" * 60)

    state = load_state()
    state["last_check"] = stamp_now()
    save_state(state)

    # 退出码：有变更/新增/缺失返回 1，便于脚本化
    return 1 if (changed or new or missing) else 0


def cmd_maybe(interval=None):
    state = load_state()
    if interval is not None:
        state["interval_days"] = interval
    iv = state.get("interval_days", 7)
    last = state.get("last_check")
    if not last:
        print(f"[maybe] 从未检查过，立即执行 check（间隔={iv}天）")
        return cmd_check()
    last_dt = datetime.datetime.fromisoformat(last)
    now = datetime.datetime.now()
    delta = (now - last_dt).days
    if delta >= iv:
        print(f"[maybe] 距上次检查 {delta} 天 ≥ 间隔 {iv} 天，自动执行 check：")
        return cmd_check()
    else:
        print(f"[maybe] 距上次检查 {delta} 天 < 间隔 {iv} 天，未超期，跳过。")
        print(f"[maybe] 上次检查时间：{last}")
        return 0


def main():
    import argparse
    p = argparse.ArgumentParser(description="DH_DataBase 源文件新鲜度巡检")
    sub = p.add_subparsers(dest="cmd")
    sub.add_parser("baseline", help="全量刷新指纹到 MANIFEST.xlsx")
    sub.add_parser("check", help="比对当前与基线，输出报告")
    m = sub.add_parser("maybe", help="超间隔自动 check，否则跳过")
    m.add_argument("--interval", type=int, default=None, help="覆盖间隔天数（默认取 state 文件）")
    args = p.parse_args()
    cmd = args.cmd or "maybe"
    if cmd == "baseline":
        cmd_baseline()
    elif cmd == "check":
        sys.exit(cmd_check())
    elif cmd == "maybe":
        sys.exit(cmd_maybe(args.interval))


if __name__ == "__main__":
    main()
