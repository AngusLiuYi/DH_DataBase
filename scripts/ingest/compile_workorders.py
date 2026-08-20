#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
compile_workorders.py — 华东工单提取 编译进 wiki
源: raw/fae-qa/华东工单提取仅技术服务&高价值FAQ_20260820.xlsx
产出: wiki/topic-华东工单-<产品线>.md (4 页) + wiki/topic-华东工单案例库.md (总索引)
规则(Fly 2026-08-20 决策): 全部纳入(含【禁】类); 按产品线分 4~5 页
"""
import openpyxl, hashlib, os, datetime, re

SRC = 'raw/fae-qa/华东工单提取仅技术服务&高价值FAQ_20260820.xlsx'
OUTDIR = 'wiki'

# 产品线 -> (文件名短名, 归类集合)
PL_MAP = {
    '电爪': ('电爪', {'电爪'}),
    '电缸': ('电缸', {'电缸'}),
    '音圈/直驱': ('音圈直驱', {'音圈/直驱'}),
    '驱动': ('驱动及其他', {'驱动'}),
    '通讯盒': ('驱动及其他', {'通讯盒'}),
    '柔性线': ('驱动及其他', {'柔性线'}),
    '灵巧手': ('驱动及其他', {'灵巧手'}),
    'DHEX': ('驱动及其他', {'DHEX'}),
    '智能工具': ('驱动及其他', {'智能工具'}),
}

def sha256(path):
    return hashlib.sha256(open(path, 'rb').read()).hexdigest()

def fmt_date(v):
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, str):
        m = re.match(r'(\d{4}-\d{2}-\d{2})', v)
        if m: return m.group(1)
    return str(v) if v else ''

def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb['服务工单']
    hdr = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
    col = {h: i+1 for i, h in enumerate(hdr)}

    rows = []
    seen_ids = set()
    dup = 0
    for r in range(2, ws.max_row+1):
        rec = {h: ws.cell(row=r, column=c).value for h, c in col.items()}
        sid = rec.get('服务单号')
        if sid in seen_ids:
            dup += 1
            continue
        if sid:
            seen_ids.add(sid)
        rows.append(rec)

    print(f'读取数据行(去重后): {len(rows)} (去重跳过 {dup} 条重复单号)')

    # 按产品线分组
    groups = {k: [] for k in PL_MAP}
    unmatched = []
    for rec in rows:
        pl = rec.get('执行器产品线')
        if pl in PL_MAP:
            groups[pl].append(rec)
        else:
            unmatched.append(rec)
    if unmatched:
        print(f'⚠️ 未匹配产品线: {len(unmatched)} 条, 产品线值={set(r.get("执行器产品线") for r in unmatched)}')

    # 每个产品线 -> 目标文件
    file_groups = {}
    for pl, recs in groups.items():
        target = PL_MAP[pl][0]
        file_groups.setdefault(target, []).extend(recs)

    # 总统计(用于索引)
    total = len(rows)
    pl_counts = {fname: len(recs) for fname, recs in file_groups.items()}
    # 问题类型整体分布
    type_counter = {}
    for rec in rows:
        t = rec.get('问题类型') or '未分类'
        type_counter[t] = type_counter.get(t, 0) + 1
    # 时间范围
    dts = [rec.get('实际服务时间') for rec in rows]
    def todt(v):
        if isinstance(v, datetime.datetime): return v
        if isinstance(v, str):
            try: return datetime.datetime.strptime(v[:19], '%Y-%m-%d %H:%M:%S')
            except: return None
        return None
    dts = [todt(v) for v in dts if todt(v)]
    date_min = min(dts).strftime('%Y-%m-%d') if dts else ''
    date_max = max(dts).strftime('%Y-%m-%d') if dts else ''
    customers = set(rec.get('服务客户名称') for rec in rows if rec.get('服务客户名称'))
    src_sha = sha256(SRC)
    src_size = os.path.getsize(SRC)
    src_mtime = datetime.datetime.fromtimestamp(os.path.getmtime(SRC)).strftime('%Y-%m-%dT%H:%M:%S')

    # 生成各产品线页
    page_meta = {}
    for fname, recs in file_groups.items():
        md = build_page(fname, recs, SRC, src_sha)
        out = os.path.join(OUTDIR, f'topic-华东工单-{fname}.md')
        with open(out, 'w', encoding='utf-8') as f:
            f.write(md)
        page_meta[fname] = (out, len(recs))
        print(f'生成 {out} ({len(recs)} 条)')

    # 生成总索引页
    idx = build_index(total, pl_counts, type_counter, date_min, date_max,
                      len(customers), src_sha, src_size, src_mtime, SRC, dup, page_meta)
    out = os.path.join(OUTDIR, 'topic-华东工单案例库.md')
    with open(out, 'w', encoding='utf-8') as f:
        f.write(idx)
    print(f'生成 {out} (总索引)')

    # 输出 MANIFEST 用指纹
    print('\n=== MANIFEST 登记信息 ===')
    print(f'文件名: {os.path.basename(SRC)}')
    print(f'sha256: {src_sha}')
    print(f'mtime: {src_mtime}')
    print(f'size: {src_size}')
    print(f'数据行(去重): {total}')

def build_page(fname, recs, src, src_sha):
    # 按问题类型分组
    by_type = {}
    for rec in recs:
        t = rec.get('问题类型') or '未分类'
        by_type.setdefault(t, []).append(rec)
    # 类型按数量降序
    types_sorted = sorted(by_type.items(), key=lambda x: -len(x[1]))

    lines = []
    lines.append('---')
    lines.append(f'title: 华东工单案例库 · {fname}')
    lines.append('type: topic')
    lines.append(f'created: 2026-08-20')
    lines.append(f'updated: 2026-08-20')
    lines.append(f'sources: [{os.path.basename(src)}]')
    lines.append(f'product_lines: [{fname}]')
    lines.append('tags: [工单, 案例库, 华东, 技术服务, 高价值]')
    lines.append('---')
    lines.append('')
    lines.append(f'# 华东工单案例库 · {fname}')
    lines.append('')
    lines.append(f'> **源文件**：`{src}`（华东技术服务工单提取，仅技术服务&高价值，2026-01-01~2026-08-17）')
    lines.append(f'> **归类**：[[topic-华东工单案例库]]（总索引）')
    lines.append(f'> **范围**：本页 {fname} 产品线，共 **{len(recs)}** 条工单，含【禁】类（Fly 决策：全部纳入）')
    lines.append('')
    lines.append('## 本页问题类型分布')
    lines.append('')
    for t, rs in types_sorted:
        lines.append(f'- {t}：{len(rs)}')
    lines.append('')
    lines.append('---')
    lines.append('')

    for t, rs in types_sorted:
        lines.append(f'## {t}（{len(rs)}）')
        lines.append('')
        # 按时间倒序
        rs_sorted = sorted(rs, key=lambda r: (str(r.get('实际服务时间') or ''),), reverse=True)
        for rec in rs_sorted:
            sid = rec.get('服务单号') or '?'
            cust = rec.get('服务客户名称') or ''
            model = rec.get('执行器型号1') or ''
            drv = rec.get('驱动器') or ''
            srv = rec.get('服务人员') or ''
            grp = rec.get('服务人员组别') or ''
            dt = fmt_date(rec.get('实际服务时间'))
            phen = (rec.get('问题现象描述') or '').strip().replace('\n', ' ')
            ana = (rec.get('分析步骤') or '').strip().replace('\n', ' ')
            concl = (rec.get('结论及处理') or '').strip().replace('\n', ' ')
            part = (rec.get('本次问题部件') or '').strip()
            meta_bits = []
            if cust: meta_bits.append(f'客户：{cust}')
            if model: meta_bits.append(f'型号：{model}')
            if drv: meta_bits.append(f'驱动器：{drv}')
            if srv or grp: meta_bits.append(f'服务：{srv}{(" " + grp) if grp else ""}')
            if dt: meta_bits.append(f'时间：{dt}')
            lines.append(f'### {sid}')
            if meta_bits:
                lines.append('> ' + ' ｜ '.join(meta_bits))
            if phen:
                lines.append(f'- **现象**：{phen}')
            if ana and concl and ana != concl:
                lines.append(f'- **分析**：{ana}')
                lines.append(f'- **结论**：{concl}')
            elif concl:
                lines.append(f'- **分析/结论**：{concl}')
            elif ana:
                lines.append(f'- **分析/结论**：{ana}')
            if part:
                lines.append(f'- **问题部件**：{part}')
            lines.append('')
        lines.append('---')
        lines.append('')
    return '\n'.join(lines)

def build_index(total, pl_counts, type_counter, date_min, date_max,
                n_customers, src_sha, src_size, src_mtime, src, dup, page_meta):
    lines = []
    lines.append('---')
    lines.append('title: 华东工单案例库（总索引）')
    lines.append('type: topic')
    lines.append('created: 2026-08-20')
    lines.append('updated: 2026-08-20')
    lines.append(f'sources: [{os.path.basename(src)}]')
    lines.append('product_lines: [电爪, 电缸, 音圈/直驱, 驱动]')
    lines.append('tags: [工单, 案例库, 华东, 技术服务, 高价值, 索引]')
    lines.append('---')
    lines.append('')
    lines.append('# 华东工单案例库（总索引）')
    lines.append('')
    lines.append(f'> **源文件**：`{src}`')
    lines.append(f'> **哈希指纹(sha256)**：`{src_sha}`')
    lines.append(f'> **大小/修改时间**：{src_size} 字节 / {src_mtime}')
    lines.append(f'> **范围**：华东技术服务工单提取，仅「技术服务 & 高价值 FAQ」，**全部纳入含【禁】类**（Fly 2026-08-20 决策）')
    lines.append(f'> **时间跨度**：{date_min} ~ {date_max} ｜ **数据行(去重后)**：{total}（去重跳过 {dup} 条重复单号）')
    lines.append('')
    lines.append('## 概况')
    lines.append('')
    lines.append(f'- 总工单：**{total}** 条（全部为「技术服务 / 高价值」）')
    lines.append(f'- 涉及客户：**{n_customers}** 家')
    lines.append(f'- 产品线分布（按页）：' + '、'.join(f'{k} {v}' for k, v in sorted(pl_counts.items(), key=lambda x: -x[1])))
    lines.append('')
    lines.append('## 子页（按产品线）')
    lines.append('')
    for fname in ['电爪', '电缸', '音圈直驱', '驱动及其他']:
        if fname in pl_counts:
            lines.append(f'- [[topic-华东工单-{fname}]] — {fname} 产品线，{pl_counts[fname]} 条')
    lines.append('')
    lines.append('## 问题类型整体分布（Top 20）')
    lines.append('')
    for t, c in sorted(type_counter.items(), key=lambda x: -x[1])[:20]:
        lines.append(f'- {t}：{c}')
    lines.append('')
    lines.append('## 说明')
    lines.append('')
    lines.append('- 【禁】类（客户应用使用 / 简易咨询 / 选型 / 售前 / 产品维修等）已按 Fly 决策**全部纳入**，其标签保留在条目「问题类型」中，便于追溯。')
    lines.append('- 每条工单保留：服务单号、客户、型号、驱动器、服务人员/组别、时间、现象、分析、结论、问题部件。')
    lines.append('- 源文件为只读原始资料；本索引与子页为 AI 编译层，不修改原件。')
    lines.append('')
    return '\n'.join(lines)

if __name__ == '__main__':
    main()
