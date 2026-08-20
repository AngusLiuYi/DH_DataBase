# -*- coding: utf-8 -*-
import openpyxl, html, random
from collections import defaultdict

SRC = r"D:\AI\DH_DataBase\raw\fae-qa\华东工单提取FAQ_annotated.xlsx"
OUT = r"D:\AI\DH_DataBase\build\classify_full_report.html"

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
ws = wb["服务工单"]
hdr = [c.value for c in ws[1]]
col = {n: i + 1 for i, n in enumerate(hdr)}

def g(r, n):
    return str(ws.cell(r, col[n]).value or "")

rows = []
for r in range(2, ws.max_row + 1):
    v = ws.cell(r, col["参考价值"]).value
    if v is None:
        continue
    rows.append(r)

buck = defaultdict(list)
for r in rows:
    note = g(r, "筛选备注(理由)")
    if note.startswith("LLM未重判"):
        buck["待补判(额度402)"].append(r)
    else:
        buck[g(r, "参考价值")].append(r)

cnt = {k: len(v) for k, v in buck.items()}
random.seed(2026)
for k in buck:
    random.shuffle(buck[k])

badge = {"高价值": "#2e7d32", "中价值": "#ef6c00", "舍弃": "#c62828", "待补判(额度402)": "#6a1b9a"}
order = ["高价值", "中价值", "舍弃", "待补判(额度402)"]

def esc(s):
    return html.escape(str(s))

parts = []
parts.append(f"""<!DOCTYPE html><html lang="zh"><head><meta charset="utf-8">
<style>
body{{font-family:-apple-system,'Microsoft YaHei',sans-serif;margin:0;background:#f5f5f5;color:#222}}
header{{background:#1a237e;color:#fff;padding:16px 24px}}
h1{{margin:0;font-size:20px}} .sub{{opacity:.85;font-size:13px;margin-top:4px}}
.wrap{{padding:16px 24px}}
.summary{{display:flex;gap:12px;margin:16px 0;flex-wrap:wrap}}
.card{{flex:1;min-width:140px;background:#fff;border-radius:8px;padding:14px;text-align:center;box-shadow:0 1px 3px rgba(0,0,0,.1)}}
.card .n{{font-size:28px;font-weight:700}} .card .l{{font-size:13px;color:#666}}
.sec{{margin:20px 0}} .sec h2{{font-size:16px;border-left:4px solid #1a237e;padding-left:8px}}
.item{{background:#fff;border-radius:8px;padding:12px 14px;margin:8px 0;box-shadow:0 1px 2px rgba(0,0,0,.08)}}
.item .top{{font-size:12px;color:#888;margin-bottom:6px}}
.item .lab{{display:inline-block;color:#fff;font-size:12px;padding:1px 8px;border-radius:10px;margin-left:6px}}
.kv{{font-size:13px;margin:3px 0}} .k{{color:#1565c0;font-weight:600}}
.reason{{font-size:13px;background:#fff8e1;border-left:3px solid #f9a825;padding:6px 10px;margin-top:6px;border-radius:0 4px 4px 0}}
.note{{background:#fff3e0;border:1px solid #ffb74d;border-radius:8px;padding:12px 16px;margin:16px 0;font-size:14px;line-height:1.6}}
</style></head><body>
<header><h1>华东工单 LLM 语义重筛 · 全量报告</h1>
<div class="sub">模型 deepseek-v4-flash | 来源 华东工单提取FAQ.xlsx | 全量 4905 条 | 生成于 2026-08-20</div></header>
<div class="wrap">
<div class="note"><b>重要说明（数据可信度）：</b>全量 4905 条中，<b>4867 条</b>已由 deepseek-v4-flash 逐条语义重判（高 {cnt.get('高价值',0)} / 中 {cnt.get('中价值',0)} / 舍 {cnt.get('舍弃',0)}）。
另有 <b>{cnt.get('待补判(额度402)',0)} 条</b>因 <b>DeepSeek key 额度耗尽（HTTP 402 Payment Required）</b>未能完成 LLM 重判，沿用旧规则分档值（35 高 + 3 中），备注已标「LLM未重判(额度402回退旧值)」，请在 xlsx 中据此筛选复核。充值后可单独补跑这 {cnt.get('待补判(额度402)',0)} 条。</div>
<div class="summary">
<div class="card"><div class="n" style="color:{badge['高价值']}">{cnt.get('高价值',0)}</div><div class="l">高价值</div></div>
<div class="card"><div class="n" style="color:{badge['中价值']}">{cnt.get('中价值',0)}</div><div class="l">中价值</div></div>
<div class="card"><div class="n" style="color:{badge['舍弃']}">{cnt.get('舍弃',0)}</div><div class="l">舍弃（额外剔除）</div></div>
<div class="card"><div class="n" style="color:{badge['待补判(额度402)']}">{cnt.get('待补判(额度402)',0)}</div><div class="l">待补判(额度402)</div></div>
</div>""")

SAMP = 18
for lab in order:
    items = buck.get(lab, [])
    parts.append(f'<div class="sec"><h2>{lab}（共 {len(items)} 条，以下抽样 {min(SAMP,len(items))} 条）</h2>')
    for r in items[:SAMP]:
        parts.append(f'''<div class="item">
<div class="top">单号 {esc(g(r,'服务单号'))} · 行号 {r} · <span class="lab" style="background:{badge[lab]}">{lab}</span></div>
<div class="kv"><span class="k">类型</span> {esc(g(r,'问题类型'))}</div>
<div class="kv"><span class="k">现象</span> {esc(g(r,'问题现象描述'))}</div>
<div class="kv"><span class="k">步骤</span> {esc(g(r,'分析步骤'))}</div>
<div class="kv"><span class="k">结论</span> {esc(g(r,'结论及处理'))}</div>
<div class="reason"><b>LLM理由：</b>{esc(g(r,'筛选备注(理由)'))}</div>
</div>''')
    parts.append('</div>')

parts.append('</div></body></html>')
open(OUT, "w", encoding="utf-8").write("\n".join(parts))
print("REPORT_WRITTEN", OUT, "| total", len(rows), "| cnt", dict(cnt))
