#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""华东工单语义重筛：调用 DeepSeek API 逐条判别 高/中/舍弃。
读取 华东工单提取FAQ.xlsx 中现有 参考价值∈{高价值,中价值} 的行（约4905条），
逐条用模型按语义重新判定，写回 参考价值 与 筛选备注(理由)。
判别标准重在语义（长但水/套话/伪答案/流水账降档），不靠字数。
"""
import openpyxl, json, os, sys, time, argparse, urllib.request, urllib.error
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import Counter

TGT = r"D:\AI\DH_DataBase\raw\fae-qa\华东工单提取FAQ.xlsx"
BASE = "https://api.deepseek.com"
MODEL = "deepseek-v4-flash"
ENV_PATH = r"D:\AI\DH_DataBase\build\configs\deepseek.env"
OUT_JSON = r"D:\AI\DH_DataBase\build\classify_deepseek_result.json"
CONCURRENCY = 8
MAX_RETRIES = 4

SYSTEM = """你是一名资深 FAE（现场应用工程师）售后工单质检专家。你的任务是判断一条已闭环/处理的售后工单，是否值得作为经验沉淀进 FAQ/知识库。
请按以下语义标准，而不是字数多少，来判定：

【高价值】——信息真实、可作为经验或 FAQ 录入：
- 咨询/规格类（客户问用法、参数、选型）：问题明确，且答案是实打实的技术信息（如具体参数、动作含义、限制条件），不是"已告知客户""详见手册"这类套话。
- 故障/调试类：现象描述具体（出什么错、什么条件下出现），且分析步骤体现了真实排查动作（测量、替换对比、参数调整、定位、核对寄存器/接线等），结论给出了明确的处置与结果。
- 即使未闭环，只要含有完整、可复用的技术内容，也可判高价值。

【中价值】——有内容但需人工审核/完善：
- 某一栏偏弱或残缺（如现象具体但结论只写"已处理"没说怎么处理；或步骤是流水账但能看出干了啥）；
- 未闭环但有实质待跟进；
- 培训/SOP 类资料。

【舍弃】——纯噪声，直接丢弃：
- 现象与结论都无有效内容（如"演示""测试正常""无"）；
- 流水账且无技术信息（如"1.现场 2.沟通 3.跟进"却没有诊断动作和结果）；
- 套话结论无实质（如"已处理""正常""客户陪产"且无任何技术细节）；
- 陪产/现场配合类记录。

只输出 JSON，格式：{"label":"高价值|中价值|舍弃","reason":"一句话说明判定依据"}。
务必基于语义判断，不要因为文字长就给高价值，也不要因为短就给低价值。"""

USER_TMPL = """【问题类型】{typ}
【问题现象】{phen}
【分析步骤】{steps}
【结论及处理】{concl}

请判定该工单属于 高价值 / 中价值 / 舍弃，并给出理由。"""


def load_key():
    k = os.environ.get("DEEPSEEK_API_KEY")
    if k:
        return k.strip()
    if os.path.exists(ENV_PATH):
        for line in open(ENV_PATH, encoding="utf-8"):
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                kk, vv = line.split("=", 1)
                if kk.strip() == "DEEPSEEK_API_KEY":
                    return vv.strip().strip('"').strip("'")
    return None


def chat(key, typ, phen, steps, concl, retries=MAX_RETRIES):
    url = BASE + "/v1/chat/completions"
    data = {
        "model": MODEL,
        "messages": [
            {"role": "system", "content": SYSTEM},
            {"role": "user", "content": USER_TMPL.format(typ=typ, phen=phen, steps=steps, concl=concl)},
        ],
        "response_format": {"type": "json_object"},
        "temperature": 0,
        "max_tokens": 300,
    }
    body = json.dumps(data).encode("utf-8")
    last_err = ""
    for attempt in range(1, retries + 1):
        try:
            req = urllib.request.Request(
                url, data=body,
                headers={"Content-Type": "application/json", "Authorization": "Bearer " + key},
                method="POST",
            )
            with urllib.request.urlopen(req, timeout=60) as resp:
                obj = json.loads(resp.read().decode("utf-8"))
            content = obj["choices"][0]["message"]["content"]
            parsed = json.loads(content)
            label = str(parsed.get("label", "")).strip()
            reason = str(parsed.get("reason", "")).strip()
            if label in ("高价值", "中价值", "舍弃"):
                return label, reason
            return "中价值", "模型label异常:" + label + " | " + reason
        except Exception as e:
            last_err = str(e)
            if attempt == retries:
                return "ERROR", "API失败:" + last_err
            time.sleep(2 * attempt)
    return "ERROR", "unknown:" + last_err


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=0, help="只跑前 N 条（测试用），0=全量")
    ap.add_argument("--no-write", action="store_true", help="只出结果不写回 xlsx")
    ap.add_argument("--offset", type=int, default=0, help="从第 N 条（0-based）开始取")
    ap.add_argument("--resume", action="store_true", help="从已有结果JSON续跑，跳过已完成行")
    args = ap.parse_args()

    key = load_key()
    if not key:
        print("ERROR: 未找到 DEEPSEEK_API_KEY（环境变量或 %s）" % ENV_PATH)
        sys.exit(2)

    # key 验证：先用一条 dummy 数据探一次
    v_label, v_reason = chat(key, "测试", "测试现象", "测试步骤", "测试结论")
    if v_label == "ERROR":
        print("KEY 验证失败:", v_reason)
        sys.exit(4)
    print("KEY 验证通过（模型返回示例 label=%s）" % v_label)

    wb = openpyxl.load_workbook(TGT)
    ws = wb["服务工单"]
    hdr = [c.value for c in ws[1]]
    col = {n: i + 1 for i, n in enumerate(hdr)}
    need = ["问题类型", "问题现象描述", "分析步骤", "结论及处理", "参考价值", "筛选备注(理由)", "服务单号"]
    for n in need:
        if n not in col:
            print("缺少列:", n)
            sys.exit(3)

    rows = [r for r in range(2, ws.max_row + 1)
            if ws.cell(r, col["参考价值"]).value in ("高价值", "中价值")]
    print("待重判总行数:", len(rows))
    if args.offset:
        rows = rows[args.offset:]
    if args.limit:
        rows = rows[:args.limit]
        print("本次运行条数(limit=%d):" % args.limit, len(rows))
    else:
        print("本次运行条数(全量):", len(rows))

    out = OUT_JSON.replace(".json", "_dryrun.json") if args.no_write else OUT_JSON
    results = {}
    # 断点续跑：载入已完成行
    if args.resume and os.path.exists(out):
        try:
            prev = json.load(open(out, encoding="utf-8"))
            for d in prev:
                results[d["row"]] = (d["no"], d["label"], d["reason"])
            print("续跑已载入已完成:", len(results))
        except Exception as e:
            print("续跑载入失败，重新开始:", e)
    rows = [r for r in rows if r not in results]
    print("本次需跑行数:", len(rows))

    def work(r):
        typ = ws.cell(r, col["问题类型"]).value or ""
        phen = ws.cell(r, col["问题现象描述"]).value or ""
        steps = ws.cell(r, col["分析步骤"]).value or ""
        concl = ws.cell(r, col["结论及处理"]).value or ""
        no = ws.cell(r, col["服务单号"]).value or ""
        label, reason = chat(key, str(typ), str(phen), str(steps), str(concl))
        return r, no, label, reason

    done = 0
    errs = 0
    with ThreadPoolExecutor(max_workers=CONCURRENCY) as ex:
        futs = {ex.submit(work, r): r for r in rows}
        for f in as_completed(futs):
            r, no, label, reason = f.result()
            results[r] = (no, label, reason)
            done += 1
            if label == "ERROR":
                errs += 1
            if done % 50 == 0:  # checkpoint
                json.dump(
                    [{"row": rr, "no": v[0], "label": v[1], "reason": v[2]} for rr, v in results.items()],
                    open(out, "w", encoding="utf-8"), ensure_ascii=False, indent=1,
                )
                print(f"进度 {done}/{len(rows)} (err={errs}) [checkpoint]")
            elif done % 100 == 0:
                print(f"进度 {done}/{len(rows)} (err={errs})")

    cnt = Counter(v[1] for v in results.values())
    print("分档统计:", dict(cnt), " 错误行:", errs)

    # 写回（非 dry-run 时）
    if args.no_write:
        print("[DRY-RUN] 未写回 xlsx")
    else:
        for r, (no, label, reason) in results.items():
            if label == "ERROR":
                # 保留原参考价值，仅在理由列记录错误，避免污染数据
                ws.cell(r, col["筛选备注(理由)"]).value = "LLM-ERROR:" + reason
                continue
            ws.cell(r, col["参考价值"]).value = label
            ws.cell(r, col["筛选备注(理由)"]).value = "LLM:" + reason
        wb.save(TGT)
        print("已写回:", TGT)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(
            [{"row": r, "no": v[0], "label": v[1], "reason": v[2]} for r, v in results.items()],
            f, ensure_ascii=False, indent=1,
        )
    print("日志:", out)


if __name__ == "__main__":
    main()
