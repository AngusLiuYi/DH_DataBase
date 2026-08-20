# -*- coding: utf-8 -*-
"""
把「华东工单提取仅技术服务&高价值FAQ_20260820.xlsx」的 2772 条高价值技术服务工单
逐条投喂进 FastGPT 数据集（作为知识库语料）。

设计：
- 直接读 xlsx 源，不依赖 markdown 解析
- 每条拼成结构化文本：类型/现象/排查步骤/结论/客户/型号/部件
- 先验证连通性（list datasets），再 create/reuse 数据集，再逐条 insert text
- 并发 + 限流 + 失败重试 + 每 50 条 checkpoint（可中断续跑）

凭据来源（二选一，不进 git）：
- build/configs/kb.env  写 KbBaseUrl=http://xxx  KbApiKey=fastgpt-xxx
- 或环境变量 KB_BASE_URL / KB_API_KEY

用法：
  python feed_fastgpt.py            # 全量投喂
  python feed_fastgpt.py --limit 20 # 先投 20 条试水
  python feed_fastgpt.py --resume   # 从 checkpoint 续跑
"""
import openpyxl, json, os, sys, time, urllib.request, urllib.error
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import Counter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_PATH = os.path.normpath(os.path.join(SCRIPT_DIR, "..", "build", "configs", "kb.env"))
SRC = r"D:\AI\DH_DataBase\raw\fae-qa\华东工单提取仅技术服务&高价值FAQ_20260820.xlsx"
DATASET_NAME = "华东工单FAQ-技术服务高价值"
CONCURRENCY = 4
CKPT_JSON = r"D:\AI\DH_DataBase\build\feed_fastgpt_result.json"

# FastGPT 开放 API 路径（v4/v5 通用，若版本不同按实际响应微调）
P_LIST = "/api/core/dataset/list"
P_CREATE = "/api/core/dataset/create"
P_INSERT = "/api/core/dataset/collection/create/text"


def load_kb():
    base, key = "", ""
    if os.path.exists(ENV_PATH):
        for line in open(ENV_PATH, encoding="utf-8"):
            line = line.strip()
            if line.startswith("KbBaseUrl="):
                base = line.split("=", 1)[1].strip().rstrip("/")
            elif line.startswith("KbApiKey="):
                key = line.split("=", 1)[1].strip()
    base = os.environ.get("KB_BASE_URL", base)
    key = os.environ.get("KB_API_KEY", key)
    return base, key


def req(base, key, method, path, body=None, retries=3):
    url = base + path
    data = json.dumps(body).encode("utf-8") if body is not None else None
    last = None
    for i in range(retries):
        r = urllib.request.Request(url, data=data, method=method)
        r.add_header("Authorization", "Bearer " + key)
        r.add_header("Content-Type", "application/json")
        try:
            with urllib.request.urlopen(r, timeout=40) as resp:
                return resp.getcode(), json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            last = (e.code, e.read().decode("utf-8", "ignore"))
            if e.code in (401, 403, 404):  # 凭据/路径错，不重试
                return last
            time.sleep(2 * (i + 1))
        except Exception as e:
            last = (-1, str(e))
            time.sleep(2 * (i + 1))
    return last if last else (-1, "unknown")


def build_text(row):
    def g(name):
        return str(row.get(name, "") or "").strip()
    typ = g("问题类型"); phen = g("问题现象描述"); steps = g("分析步骤")
    concl = g("结论及处理"); cust = g("服务客户名称"); model = g("执行器型号1")
    part = g("本次问题部件"); no = g("服务单号")
    parts = []
    if typ:
        parts.append("【%s】" % typ)
    if phen:
        parts.append("现象：%s" % phen)
    if steps:
        parts.append("排查步骤：%s" % steps)
    if concl:
        parts.append("结论及处理：%s" % concl)
    meta = []
    if cust:
        meta.append("客户:%s" % cust)
    if model:
        meta.append("型号:%s" % model)
    if part:
        meta.append("部件:%s" % part)
    if no:
        meta.append("单号:%s" % no)
    if meta:
        parts.append("（%s）" % " ".join(meta))
    return "\n".join(parts)


def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--resume", action="store_true")
    args = ap.parse_args()

    base, key = load_kb()
    if not base or not key:
        print("ERROR: 未找到 KbBaseUrl/KbApiKey（configs/kb.env 或环境变量 KB_BASE_URL/KB_API_KEY）")
        sys.exit(2)

    # 1) 连通性验证
    code, resp = req(base, key, "GET", P_LIST)
    print("[连通性] %s -> %s" % (P_LIST, code))
    if code != 200:
        print("  !! 连通失败，请检查 KbBaseUrl / KbApiKey / 网络。响应:", str(resp)[:300])
        sys.exit(4)
    print("  连通成功。响应(前200):", str(resp)[:200])

    # 2) 建/复用数据集
    dataset_id = None
    # list 返回里找同名
    try:
        dlist = resp.get("data", {}).get("datasets", []) if isinstance(resp, dict) else []
    except Exception:
        dlist = []
    for d in dlist:
        if d.get("name") == DATASET_NAME:
            dataset_id = d.get("id") or d.get("_id")
            print("[数据集] 复用已存在 id=%s" % dataset_id)
            break
    if not dataset_id:
        c, cr = req(base, key, "POST", P_CREATE, {"name": DATASET_NAME})
        print("[数据集] 创建 ->", c, str(cr)[:200])
        if c == 200 and isinstance(cr, dict):
            dataset_id = cr.get("data", {}).get("id") or cr.get("data", {}).get("_id")
        if not dataset_id:
            print("  !! 创建数据集失败，停止。")
            sys.exit(5)
        print("[数据集] 新建 id=%s" % dataset_id)

    # 3) 读源
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["服务工单"]
    hdr = [c.value for c in ws[1]]
    col = {n: i + 1 for i, n in enumerate(hdr)}
    rows = []
    for r in range(2, ws.max_row + 1):
        d = {n: ws.cell(r, col[n]).value for n in col}
        if not (d.get("问题现象描述") or d.get("结论及处理")):
            continue
        rows.append(d)
    print("[源] 有效工单行数:", len(rows))
    if args.limit:
        rows = rows[:args.limit]
        print("[源] limit=%d 取前 %d 条" % (args.limit, len(rows)))

    # 4) 断点续跑
    done = set()
    if args.resume and os.path.exists(CKPT_JSON):
        for x in json.load(open(CKPT_JSON, encoding="utf-8")):
            if x.get("ok"):
                done.add(x["no"])
        print("[续跑] 已完成 %d 条" % len(done))

    batch = [d for d in rows if d.get("服务单号") not in done]
    print("[待投] %d 条" % len(batch))

    results = []
    done_n = 0
    err_n = 0

    def work(d):
        text = build_text(d)
        no = d.get("服务单号")
        c, cr = req(base, key, "POST", P_INSERT, {"datasetId": dataset_id, "text": text})
        ok = (c == 200)
        reason = "" if ok else str(cr)[:120]
        return no, ok, reason

    with ThreadPoolExecutor(max_workers=CONCURRENCY) as ex:
        futs = {ex.submit(work, d): d for d in batch}
        for f in as_completed(futs):
            no, ok, reason = f.result()
            results.append({"no": no, "ok": ok, "reason": reason})
            done_n += 1
            if not ok:
                err_n += 1
            if done_n % 50 == 0:
                json.dump(results, open(CKPT_JSON, "w", encoding="utf-8"),
                          ensure_ascii=False, indent=1)
                print("进度 %d/%d (err=%d)" % (done_n, len(batch), err_n))

    json.dump(results, open(CKPT_JSON, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    cnt = Counter("OK" if x["ok"] else "FAIL" for x in results)
    print("[完成] 本次 %d 条 -> %s | 失败 %d" % (len(results), dict(cnt), err_n))
    if err_n:
        print("失败样例:")
        for x in [r for r in results if not r["ok"]][:5]:
            print("  ", x["no"], x["reason"][:80])


if __name__ == "__main__":
    main()
