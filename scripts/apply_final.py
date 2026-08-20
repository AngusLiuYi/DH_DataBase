#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""把 classify_deepseek_result.json 的分类结果写回 xlsx，并补齐漏跑/失败的 60 条。
步骤：备份 -> 写回 JSON 已有(4900) -> save -> 统计 -> 补跑 ERROR+缺失(约60) -> save -> 统计。
"""
import openpyxl, json, shutil, os, sys, time
from collections import Counter
sys.path.insert(0, r"D:\AI\DH_DataBase\scripts")
from classify_deepseek import load_key, chat, TGT

BAK = TGT + ".bak_20260820"
JSON = r"D:\AI\DH_DataBase\build\classify_deepseek_result.json"

# 1. 备份
if not os.path.exists(BAK):
    shutil.copy(TGT, BAK)
    print("已备份 ->", BAK)
else:
    print("备份已存在，跳过:", BAK)

# 2. 读 JSON
results = json.load(open(JSON, encoding="utf-8"))
print("JSON 条数:", len(results))

# 3. 写回 JSON 已有行（有效写新值；ERROR 仅标记理由，保留原参考价值）
wb = openpyxl.load_workbook(TGT)
ws = wb["服务工单"]
hdr = [c.value for c in ws[1]]
col = {n: i + 1 for i, n in enumerate(hdr)}
json_rows = set()
todo = []  # 需补跑的行（ERROR 或 缺失）
for d in results:
    r = d["row"]; label = d["label"]; reason = d["reason"]
    json_rows.add(r)
    if label in ("高价值", "中价值", "舍弃"):
        ws.cell(r, col["参考价值"]).value = label
        ws.cell(r, col["筛选备注(理由)"]).value = "LLM:" + reason
    else:  # ERROR
        ws.cell(r, col["筛选备注(理由)"]).value = "LLM-ERROR:" + reason
        todo.append(r)

# 缺失行（xlsx 高/中价值但不在 JSON 中）
for r in range(2, ws.max_row + 1):
    if r in json_rows:
        continue
    v = ws.cell(r, col["参考价值"]).value
    if v in ("高价值", "中价值"):
        todo.append(r)
print("待补跑行(ERROR+缺失):", len(todo))

# 4. 先 save 核心 4900 条
wb.save(TGT)
print("核心写回完成，已 save xlsx")

# 5. 统计当前
def stat():
    c = Counter()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col["参考价值"]).value
        if v in ("高价值", "中价值", "舍弃"):
            c[v] += 1
    return c
print("核心分档:", dict(stat()))

# 6. 补跑 todo
if todo:
    key = load_key()
    print("开始补跑", len(todo), "条...")
    done = 0
    for r in todo:
        typ = str(ws.cell(r, col["问题类型"]).value or "")
        phen = str(ws.cell(r, col["问题现象描述"]).value or "")
        steps = str(ws.cell(r, col["分析步骤"]).value or "")
        concl = str(ws.cell(r, col["结论及处理"]).value or "")
        label, reason = chat(key, typ, phen, steps, concl)
        if label in ("高价值", "中价值", "舍弃"):
            ws.cell(r, col["参考价值"]).value = label
            ws.cell(r, col["筛选备注(理由)"]).value = "LLM:" + reason
        else:
            ws.cell(r, col["筛选备注(理由)"]).value = "LLM-ERROR:" + reason
        done += 1
        if done % 10 == 0:
            print(f"补跑进度 {done}/{len(todo)}")
    wb.save(TGT)
    print("补跑完成，已 save xlsx")

# 7. 最终统计
c = stat()
errs = sum(1 for r in range(2, ws.max_row + 1)
           if isinstance(ws.cell(r, col["筛选备注(理由)"]).value, str)
           and ws.cell(r, col["筛选备注(理由)"]).value.startswith("LLM-ERROR"))
print("最终分档:", dict(c), " | 仍 ERROR 标记行:", errs)
print("DONE")
