# -*- coding: utf-8 -*-
"""对 raw/fae-qa 下的技术支持 FAQ 做主题聚类，输出 TOP N 类问题。

口径（Fly 2026-08-31 确认）：
  1. 华东工单表：**全量 2772 条全部计入**。
     注：「问题类型」列带【禁】前缀的类别（客户应用使用/简易咨询/产品选型等）只是后期分类体系
     细化后停止新登记使用，**数据本身有效，不得剔除**。
  2. 只用「问题现象描述」做关键词主题聚类（不掺分析/结论，避免"重新初始化"之类污染分类）。
  3. 不输出客户维度。
  4. 噗元 FAQ 表：对「问题（提问者填写）」做同样的聚类，作为交叉验证。
分类为优先级单标签（命中即停）。
"""
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
EAST = ROOT / 'raw' / 'fae-qa' / '华东工单提取仅技术服务&高价值FAQ_20260820.xlsx'
PUYUAN = ROOT / 'raw' / 'fae-qa' / '噗元FAQ_20250424.xlsx'

# (主题名, [关键词...])  顺序 = 优先级，越靠前越优先命中
RULES = [
    ('初始化/回零/寻相失败', [
        '初始化', '回零', '回原', '找原', '寻相', '零点', '零位', '原点', '标定行程', '校准行程',
        '重新标定', '零电角', '复位', '找零', '归零']),
    ('固件/软件BUG/升级', [
        '固件', '升级', '烧录', '刷机', '软件版本', '程序版本', '软件bug', '程序bug', '版本过',
        '版本更新', '刷程序']),
    ('报警代码/报错停机', [
        '报警', '报错', '故障码', '错误码', 'er.', 'err', 'al0', 'al.', '0x8', '红灯', '故障灯',
        '停机', '停止工作', '无法工作', '报故障', '跳故障', '指示灯异常', '指示灯不', '指示灯闪']),
    ('EtherCAT/总线通讯异常', [
        'ethercat', 'ecat', '总线', '主站', '从站', '扫描不到', '扫描不出', '搜不到', '找不到设备',
        'op状态', '进入op', 'pdo', 'sdo', 'esi', 'xml', '同步周期', '丢帧', '通讯周期', '通讯盒',
        '网关', '掉站', '断连', '通讯中断', '通讯不上', '通讯异常', '通讯失败', '无法通讯',
        '不能通讯', '通讯不稳定', '通讯正常', '通讯错误']),
    ('串口/485/Modbus通讯', [
        '485', 'modbus', 'rtu', '串口', '波特率', '站号', 'rs485']),
    ('上位机/软件连接不上', [
        '连接不上', '连不上', '连接失败', '无法连接', '不能连接', '软件连', '掉线', '断开连接',
        '连接中断', '无法扫描']),
    ('力控/压力/夹持力异常', [
        '力控', '压力', '保压', '柔性', '力值', '出力', '推力', '夹持力', '夹紧力', '加持力',
        '力传感', '称重', 'nf2', '力反馈', '推压', '恒力', '力度', '夹持', '夹紧', '夹不紧',
        '拧不紧', '软着陆', '力矩']),
    ('位置偏差/位置丢失', [
        '位置偏', '位置不准', '位置偏差', '位置丢失', '位置不对', '位置错误', '跑位', '重复定位',
        '定位精度', '精度', '不到位', '坐标偏', '漂移', '偏差', '高低差', '不持平', '尺寸偏',
        '误差', '分辨率']),
    ('抖动/异响/噪音', [
        '抖动', '震动', '振动', '异响', '噪音', '噪声', '啸叫', '共振', '晃动', '声音大',
        '声音比较', '太软', '整定']),
    ('堵转/过载/过流', [
        '堵转', '过载', '过流', '电流大', '电流过', '卡死', '顶死', '卡住', '憋死',
        '推不动', '带不动', '超负荷', '负载过', '顶不动', '过压']),
    ('使能/抱闸/飞车失控', [
        '使能', '飞车', '失控', '抱闸', '掉使能', '断使能', '不保持', '下滑', '掉落', '保持力']),
    ('速度/节拍/性能测试', [
        '节拍', '速度', '跟不上', '达不到', '超时', '效率低', '加速度', '减速', '运行时间',
        '循环时间', 'cd时间', '性能测试', '产能', '效率', '过冲']),
    ('参数/寄存器/模式配置', [
        '寄存器', '0x', '增益', 'pid', '参数', '模式', '导入', '导出', '阈值', '偏置值',
        '配置', '设置', '写入']),
    ('IO/信号/到位反馈', [
        'io', 'npn', 'pnp', '数字量', '模拟量', '信号', '触发', '指令', '到位']),
    ('PLC/控制卡/主站适配', [
        'plc', '欧姆龙', '基恩士', '汇川', '固高', '凌臣', '三菱', '西门子', '控制卡',
        '板卡', '卓岚', '适配', '兼容', '工控机']),
    ('接线/供电/线缆', [
        '接线', '24v', '供电', '电源', '线缆', '断线', '接触不良', '短路', '接地', '插头',
        '端子', '线序', '缺相', '电压', '线材', '网线', '通讯线', '线']),
    ('电机/编码器/硬件损坏', [
        '编码器', '不转', '烧毁', '烧坏', '霍尔', '轴承', '丝杠', '损坏', '更换电机', '硬件故障',
        'pcb', '更换']),
    ('发热/温升', ['发热', '发烫', '温度高', '温升', '烫手', '过热', '温度']),
    ('安装/机械结构/形变', [
        '安装', '机械', '背隙', '间隙', '配合', '法兰', '偏心', '卡滞', '同心度', '垂直度',
        '形变', '变形', '断裂', '磨损', '松动']),
    ('掉电/参数保存/数据丢失', [
        '掉电', '断电', '保存', '参数丢失', '恢复出厂', '上电丢失', '重启']),
    ('无法控制/不动作/运行异常', [
        '无法控制', '不能控制', '控制不了', '无法正常', '无法使用', '用不了', '无数据反馈',
        '不动作', '无动作', '不运行', '无法运行', '运行异常', '动作异常', '卡顿', '走走停停',
        '走不到', '撑不开', '合不拢', '打滑', '不动', '不工作', '停止', '异常']),
    ('现场调试/技术支持/定制需求', [
        '现场协助', '协助', '需要技术', '需求', '要求', '定制', '调试', '培训', '赋能', '核查',
        '支持']),
    ('选型/替换/规格咨询', [
        '选型', '替换', '替代', '规格', '能否', '是否可以', '可不可以', '有没有', '是否支持',
        '推荐', '咨询', '询问', '了解', '确认', '怎么', '如何', '什么', '意思']),
]


def classify(text: str):
    """返回主标签（优先级最高的命中项），None 表示未分类"""
    t = (text or '').lower()
    for name, kws in RULES:
        for kw in kws:
            if kw in t:
                return name
    return None


def load_east():
    wb = openpyxl.load_workbook(EAST, read_only=True, data_only=True)
    rows = list(wb['服务工单'].iter_rows(values_only=True))
    wb.close()
    hdr = [str(c) if c else '' for c in rows[0]]
    idx = {h: i for i, h in enumerate(hdr)}
    data = [r for r in rows[1:] if r and r[idx['服务单号']]]
    return hdr, idx, data


def load_puyuan():
    wb = openpyxl.load_workbook(PUYUAN, read_only=True, data_only=True)
    rows = list(wb['HDFAQ'].iter_rows(values_only=True))
    wb.close()
    hdr = [str(c) if c else '' for c in rows[0]]
    idx = {h: i for i, h in enumerate(hdr)}
    data = [r for r in rows[1:] if r and r[idx['问题编号']]]
    return hdr, idx, data


def report_east():
    hdr, idx, data = load_east()
    tech = data  # 【禁】类数据有效，全量计入
    banned = [r for r in data if str(r[idx['问题类型']] or '').startswith('【禁】')]
    print('=' * 78)
    print('数据源1：华东工单提取（仅技术服务&高价值）')
    print('  统计基数：全量 %d 条（其中【禁】标记 %d 条，数据有效，计入统计）'
          % (len(data), len(banned)))
    ts = [str(r[idx['实际服务时间']])[:10] for r in tech if r[idx['实际服务时间']]]
    print('  时间跨度：%s ~ %s' % (min(ts), max(ts)))

    print('\n-- 官方【问题类型】字段分布（全量，共%d类）--'
          % len(set(str(r[idx['问题类型']]) for r in tech)))
    for k, v in Counter(str(r[idx['问题类型']]) for r in tech).most_common():
        print('  %5d  %5.1f%%  %s' % (v, v * 100 / len(tech), k))

    main = Counter()
    by_topic = defaultdict(list)
    unclassified = []
    for r in tech:
        text = str(r[idx['问题现象描述']] or '')
        m = classify(text)
        if m is None:
            unclassified.append(r)
        else:
            main[m] += 1
            by_topic[m].append(r)

    print('\n-- 主题聚类 TOP20（主标签，合计100%）--')
    print('  %-26s %6s %7s %8s' % ('问题类别', '条数', '占比', '未命中'))
    for i, (k, v) in enumerate(main.most_common(20), 1):
        print('  %2d. %-26s %5d %6.1f%%' % (i, k, v, v * 100 / len(tech)))
    print('  未分类：%d 条（%.1f%%）' % (len(unclassified), len(unclassified) * 100 / len(tech)))

    print('\n-- TOP20 类 × 产品线分布 --')
    for k, v in main.most_common(20):
        c = Counter(str(r[idx['执行器产品线']] or '(空)') for r in by_topic[k])
        print('  %-26s %s' % (k, ' / '.join('%s×%d' % (a, b) for a, b in c.most_common(4))))

    print('\n-- TOP20 类 × 根因归属（本次问题部件）--')
    for k, v in main.most_common(20):
        c = Counter(str(r[idx['本次问题部件']] or '(空)') for r in by_topic[k])
        print('  %-26s %s' % (k, ' / '.join('%s×%d' % (a, b) for a, b in c.most_common(3))))

    print('\n-- TOP20 类代表工单（每类1条）--')
    for i, (k, v) in enumerate(main.most_common(20), 1):
        r = None
        for cand in by_topic[k]:
            d = str(cand[idx['问题现象描述']] or '')
            if len(d) > 20:
                r = cand
                break
        r = r or by_topic[k][0]
        print('\n[%d] %s（%d条，%.1f%%）' % (i, k, v, v * 100 / len(tech)))
        print('   现象：%s' % str(r[idx['问题现象描述']] or '')[:120].replace('\n', ' '))
        print('   结论：%s' % str(r[idx['结论及处理']] or '')[:120].replace('\n', ' '))
        print('   根因：%s｜产品线：%s｜型号：%s' % (
            r[idx['本次问题部件']], r[idx['执行器产品线']], r[idx['执行器型号1']]))

    print('\n-- 未分类样本（前20）--')
    for r in unclassified[:20]:
        print('   -', str(r[idx['问题现象描述']] or '')[:70].replace('\n', ' '))
    return main, len(tech)


def report_puyuan():
    hdr, idx, data = load_puyuan()
    print('\n' + '=' * 78)
    print('数据源2：噗元FAQ（%d 条）' % len(data))
    for col in ['问题分类', '问题类型（可自行添加）', '主站品牌~一级分支（可自行增加）', '产品线-提问']:
        print('\n-- %s --' % col)
        for k, v in Counter(str(r[idx[col]] or '(空)') for r in data).most_common(10):
            print('  %4d  %5.1f%%  %s' % (v, v * 100 / len(data), k))
    main = Counter()
    by_topic = defaultdict(list)
    for r in data:
        text = str(r[idx['问题（提问者填写）']] or '')
        m = classify(text)
        main[m or '(未分类)'] += 1
        by_topic[m or '(未分类)'].append(r)
    print('\n-- 噗元FAQ 主题聚类 --')
    for i, (k, v) in enumerate(main.most_common(20), 1):
        print('  %2d. %-26s %3d  %5.1f%%' % (i, k, v, v * 100 / len(data)))


if __name__ == '__main__':
    report_east()
    if '--all' in sys.argv:
        report_puyuan()
