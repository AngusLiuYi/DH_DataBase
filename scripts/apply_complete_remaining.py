#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""补齐 apply_final 核心写回后仍未重判的 60 条（LLM-ERROR 标记行 + 理由非 LLM: 的高/中价值缺失行）。
仅调 API 补跑，不重复写回已完成的 4845 条。"""
import openpyxl, sys, time
from collections import Counter
sys.path.insert(0, r"D:\AI\DH_DataBase\scripts")
from classify_deepseek import load_key, chat, TGT

wb = openpyxl.load_workbook(TGT)
ws = wb["服务工单"]
hdr = [c.value for c in ws[1]]
col = {n: i + 1 for i, n in enumerate(hdr)}

todo = []
for r in range(2, ws.max_row + 1):
    v = ws.cell(r, col["参考价值"]).value
    note = ws.cell(r, col["筛选备注(理由)"]).value or ""
    if isinstance(note, str) and note.startswith("LLM-ERROR:"):
        todo.append(r)
    elif v in ("高价值", "中价值") and (not isinstance(note, str) or not note.startswith("LLM:")):
        todo.append(r)
print("待补跑:", len(todo))

key = load_key()
done = 0
for r in todo:
    typ = str(ws.cell(r, col["问题类型"]).value or "")
    phen = str(ws.cell(r, col["问题现象描述"]).value or "")
    steps = str(ws.cell(r, col["分析步骤"]).value or "")
    concl = str(ws.cell(r, col["结论及处理"]).value or "")
    label, reason = chat(key, typ, phen, steps, concl)
    if label in ("高价值", "中价值", "舍弃"):
        ws.cell(r, col["参考价值"]).value = label
        ws.cell(r, col["筛选备注(理由)"]).value = "LLM:" + reason
    else:
        ws.cell(r, col["筛选备注(理由)"]).value = "LLM-ERROR:" + reason
    done += 1
    if done % 10 == 0:
        print(f"进度 {done}/{len(todo)}")

wb.save(TGT)
print("补跑完成 save")
c = Counter()
for r in range(2, ws.max_row + 1):
    v = ws.cell(r, col["参考价值"]).value
    if v in ("高价值", "中价值", "舍弃"):
        c[v] += 1
errs = sum(1 for r in range(2, ws.max_row + 1)
           if isinstance(ws.cell(r, col["筛选备注(理由)"]).value, str)
           and ws.cell(r, col["筛选备注(理由)"]).value.startswith("LLM-ERROR"))
print("最终分档:", dict(c), "| 仍 ERROR 标记:", errs)
print("DONE")
